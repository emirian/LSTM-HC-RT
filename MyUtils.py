import scipy.io as sio
import scipy as sp
from scipy.signal import savgol_filter
from scipy import signal
import seaborn as sns
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as img

from tensorflow.keras.models import Sequential
from tensorflow.keras.models import load_model
from tensorflow.keras.layers import Dense, LSTM, Activation, Flatten, TimeDistributed, AveragePooling1D
from tensorflow.keras.callbacks import EarlyStopping, ModelCheckpoint
from tensorflow.keras.wrappers.scikit_learn import KerasRegressor
from tensorflow.keras import optimizers


from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score 
from sklearn.linear_model import Lasso
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import KFold
import openpyxl
import mne.viz as mv
from scipy.fftpack import rfft, irfft

#import tensorflow as tf
from tensorflow.keras import backend as K

def subband2Index (subband):
    switcher = {
            'bandAlpha' : 2,
            'bandAlphaFiltered':3,
            'bandBeta': 4,
            'bandBetaFiltered':5,
            'bandDelta':6,
            'bandDeltaFiltered': 7,
            'bandGamma': 8,
            'bandGammaFiltered': 9,
            'bandTheta': 10,
            'bandThetaFiltered': 11,
            'full': 12,
            }
    return switcher.get(subband)

def TrainTest(SubjectID2Exclude, subband):
    ExcCol = subband2Index(subband)
    IDs = list(range(22))
    #IDs = np.delete(IDs, SubjectID2Exclude)

    if subband == 'full':
       data  = sio.loadmat('shamdata_tlgo')
       HC    = data['shamhceeg'][:,IDs] 
    else:
 # Load the EEG data
       data  = sio.loadmat('DatahcAlireza')
       HC     = data[subband][:,IDs] # Healthy Control

    # Load the EEG data labels
    labels  = sio.loadmat('task_gvssham')
    HC_labels     = labels['hcoffmed'][:,IDs]

    fs_orig = 1000 # original sampling rate = 1000Hz
    
    ALL_EEG = HC
    ALL_Beh = HC_labels


    num_chans,time_steps,num_trials = ALL_EEG[0,1].shape
    X_ALL_EEG = np.zeros((ALL_EEG.size,num_chans,time_steps,num_trials))

    for pt in IDs: # pt=patient
        x = ALL_EEG[0,pt]
        X_ALL_EEG[pt,:,:,:] = x
    
    # Data transpose     
    X_ALL_EEG = np.transpose(X_ALL_EEG, (0, 3, 2, 1))        
    # Data dimensions          
    p,m,n,d = X_ALL_EEG.shape
    # Data reshape     
    X_ALL_EEG =  X_ALL_EEG.reshape(p* m,n,d)  


    # # # # # Load the EEG data labels # # # # # 
    num_trials, num_behvars = ALL_Beh[0,1].shape
    y_ALL = np.zeros((ALL_EEG.size, num_trials,1))            
    for pt in IDs:
        #print ("pt= " + str(pt+1))
        for tr in range (num_trials):
            i = 12 # index of Reaction Time (Change the index to look other Behavior measures)
            y_ALL[pt,tr] = ALL_Beh[0,pt][tr,i]
        
    y_ALL = np.reshape(y_ALL, (ALL_EEG.size*num_trials, 1)) 

    #------- Removing the Subject2Exclude--------------
    X_ALL_EEG_SAVED = np.delete(X_ALL_EEG, np.s_[SubjectID2Exclude*num_trials:(SubjectID2Exclude+1)*num_trials], 0)
    y_ALL_SAVED = np.delete(y_ALL, np.s_[SubjectID2Exclude*num_trials:(SubjectID2Exclude+1)*num_trials], 0)
    
    X_ALL_EEG = X_ALL_EEG_SAVED
    y_ALL = y_ALL_SAVED
    
    # # # # # Data augmentation (Down-sampling by 2) # # # # # 
    X_ALL_EEG_DS1 = X_ALL_EEG[:,0::2,:]
    X_ALL_EEG_DS2 = X_ALL_EEG[:,1::2,:]
    X_ALL_EEG_DS  = np.concatenate((X_ALL_EEG_DS1,X_ALL_EEG_DS2))
    # labels
    y_ALL_DS = np.concatenate((y_ALL,y_ALL))

    # # # # # Shuffle the data - all patients # # # # # 
    from random import shuffle
    N = X_ALL_EEG_DS.shape[0]
    indices = [i for i in range(N)]
    shuffle(indices)
    X_ALL_EEG_DS = X_ALL_EEG_DS[indices, :,:]
    y_ALL_DS     = y_ALL_DS[indices,]

    # Sanity Check: Exclude labels with NaN values
    # Indices of nan and inf values
    idx = np.where((np.isnan(y_ALL_DS)==False) & (np.isinf(y_ALL_DS)==False))
    filtered_X_ALL_EEG_DS = X_ALL_EEG_DS[idx[0],:,:]
    filtered_y_ALL_DS = y_ALL_DS[idx[0]]


    X = filtered_X_ALL_EEG_DS
    y = filtered_y_ALL_DS
    
    # Make X and y as float 32
    X = X.astype('float32')
    y = y.astype('float32')
    
    # normalize the dataset
    scaler = MinMaxScaler(feature_range=(0, 1))
    
    # Standarizing X
    m, n, d = X.shape
    X = X.reshape(m, n*d)
    X = scaler.fit_transform(X)
    X = X.reshape(m, n, d)
    # Standarizing y
    y = scaler.fit_transform(y)
    X = np.nan_to_num(X)
    # # # # # Hold-out (Training-Testing: 80%-20%) # # # # # 
    X_train,X_test,y_train,y_test=train_test_split(X,y, test_size=0.2, random_state=42)

    model = Sequential()
    model.add(LSTM(24, input_shape=(X.shape[1], X.shape[2]), return_sequences=True, implementation=2))
    model.add(TimeDistributed(Dense(1)))
    model.add(AveragePooling1D())
    model.add(Flatten())
    model.add(Dense(1, activation='linear'))

    model.compile(loss='mean_squared_error', optimizer='adam', metrics=['mean_squared_error'])
    print(model.summary())
    
    # Callback
    
    # simple early stopping
    ES = EarlyStopping(monitor='val_loss', mode='min', verbose=1, patience=500) 
    
    FileName_BestModel = 'BestModel_' + subband + '_excl' + str(SubjectID2Exclude) + '.h5'
    checkpoint_name = FileName_BestModel
    MC = ModelCheckpoint(checkpoint_name, monitor='val_loss', verbose = 1, save_best_only = True, mode ='auto')
    callbacks_list = [ES, MC]
    
    
    # Model Fitting
    history = model.fit(X_train, y_train, epochs= 1000, batch_size= 80, validation_data=(X_test, y_test), shuffle=False, callbacks=callbacks_list, verbose=1)
    
    # list all data in history
    print(history.history.keys())
    # summarize history for loss
    
    LossPlotFileName = 'Loss' + subband + '_excl' + str(SubjectID2Exclude) + '.png'

    plt.clf()
    plt.plot(history.history['loss'])
    plt.plot(history.history['val_loss'])
    plt.title('model loss')
    plt.ylabel('loss')
    plt.xlabel('epoch')
    plt.legend(['train', 'test'], loc='upper right')
    plt.savefig(LossPlotFileName)
    
    # load the saved model
    best_model = load_model(FileName_BestModel)
    
    # make predictions
    y_train_Predicted = best_model.predict(X_train)
    y_test_Predicted = best_model.predict(X_test)
    
    TrainPlotFileName = 'Train' + subband + '_excl' + str(SubjectID2Exclude) + '.png'
    TestPlotFileName  = 'Test'  + subband + '_excl' + str(SubjectID2Exclude) + '.png'
    
    # Plot Original and Predicted Train Labels
    plt.clf()
    plt.plot(y_train)
    plt.plot(y_train_Predicted)
    plt.title('Original and Predicted Train Labels based only on' + subband)
    plt.ylabel('Response Time')
    plt.xlabel('epoch')
    plt.legend(['Original', 'Predicted'], loc='upper left')
    plt.savefig(TrainPlotFileName)

    
    # Plot Original and Predicted Test Labels
    plt.clf()
    plt.plot(y_test)
    plt.plot(y_test_Predicted)
    plt.title('Original and Predicted Test Labels based only on' + subband)
    plt.ylabel('Response Time')
    plt.xlabel('epoch')
    plt.legend(['Original', 'Predicted'], loc='upper left')
    plt.savefig(TestPlotFileName)
    
    
    # Evaluate the prediction performanceof on the Standarized Data
    
    MSE_Train = mean_squared_error(y_train, y_train_Predicted)
    MSE_Test = mean_squared_error(y_test, y_test_Predicted)
    R2_score_Train = r2_score(y_train, y_train_Predicted)
    R2_score_Test = r2_score(y_test, y_test_Predicted)

    
    srcfile = openpyxl.load_workbook('Results.xlsm',read_only=False, keep_vba= True)#to open the excel sheet and if it has macros

    sheetname = srcfile.get_sheet_by_name('Train_MSE')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(MSE_Train,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    sheetname = srcfile.get_sheet_by_name('Train_R2')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(R2_score_Train,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops

    sheetname = srcfile.get_sheet_by_name('Test_MSE')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(MSE_Test,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    sheetname = srcfile.get_sheet_by_name('Test_R2')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(R2_score_Test,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    
    
    # Evaluate the prediction performanceof on the Original Data
    
    # invert predictions
    y_train_Predicted = scaler.inverse_transform(y_train_Predicted)
    y_train_Original = scaler.inverse_transform(y_train)
    
    y_test_Predicted = scaler.inverse_transform(y_test_Predicted)
    y_test_Original  = scaler.inverse_transform(y_test)
    
    
    TrainPlotFileName_T = 'T_Train' + subband + '_excl' + str(SubjectID2Exclude) + '.png'
    TestPlotFileName_T  = 'T_Test'  + subband + '_excl' + str(SubjectID2Exclude) + '.png'
    
    # Plot Original and Predicted Train Labels
    plt.clf()
    plt.plot(y_train_Original)
    plt.plot(y_train_Predicted)
    plt.title('Original and Predicted Train Labels based only on' + subband)
    plt.ylabel('Response Time')
    plt.xlabel('epoch')
    plt.legend(['Original', 'Predicted'], loc='upper left')
    plt.savefig(TrainPlotFileName_T)

    
    # Plot Original and Predicted Test Labels
    plt.clf()
    plt.plot(y_test_Original)
    plt.plot(y_test_Predicted)
    plt.title('Original and Predicted Test Labels based only on' + subband)
    plt.ylabel('Response Time')
    plt.xlabel('epoch')
    plt.legend(['Original', 'Predicted'], loc='upper left')
    plt.savefig(TestPlotFileName_T)
   
    
    
    MSE_Train = mean_squared_error(y_train_Original, y_train_Predicted)
    MSE_Test = mean_squared_error(y_test_Original, y_test_Predicted)
    
    R2_score_Train = r2_score(y_train_Original, y_train_Predicted)
    R2_score_Test = r2_score(y_test_Original, y_test_Predicted)
    
    sheetname = srcfile.get_sheet_by_name('Train_MSE_T')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(MSE_Train,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    sheetname = srcfile.get_sheet_by_name('Train_R2_T')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(R2_score_Train,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops

    sheetname = srcfile.get_sheet_by_name('Test_MSE_T')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(MSE_Test,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    sheetname = srcfile.get_sheet_by_name('Test_R2_T')
    sheetname.cell(row=SubjectID2Exclude+2,column=ExcCol).value = str(round(R2_score_Test,2)) #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
    
    srcfile.save('Results.xlsm')
      
    
#---------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------
    
def NewTest (Model, subband, PatientID):
    
    ExcCol = subband2Index(subband)
    best_model = load_model(Model)

    if subband == 'full':
       data  = sio.loadmat('shamdata_tlgo')
       HC    = data['shamhceeg'][:,PatientID] 
    else:
       data  = sio.loadmat('DatahcAlireza')
       HC     = data[subband][:,PatientID] # Healthy Control

    # Load the EEG data labels
    labels  = sio.loadmat('task_gvssham')
    HC_labels     = labels['hcoffmed'][:,PatientID]
    
    
    ALL_EEG = HC
    ALL_Beh = HC_labels


    num_chans,time_steps,num_trials = ALL_EEG[0].shape
    X_ALL_EEG = np.zeros((ALL_EEG.size,num_chans,time_steps,num_trials))
    
    
    x = ALL_EEG[0]
    X_ALL_EEG[:,:,:] = x
    # Data transpose     
    X_ALL_EEG = np.transpose(X_ALL_EEG, (0, 3, 2, 1))        
    # Data dimensions          
    p,m,n,d = X_ALL_EEG.shape
    # Data reshape     
    X_ALL_EEG =  X_ALL_EEG.reshape(p* m,n,d)  


    # # # # # Load the EEG data labels # # # # # 
    num_trials, num_behvars = ALL_Beh[0].shape
    y_ALL = np.zeros((ALL_EEG.size, num_trials,1))            

    ALLBeh = ALL_Beh[0]
    for tr in range (num_trials):
        i = 12 # index of Reaction Time (Change the index to look other Behavior measures)
        y_ALL[0,tr] = ALLBeh[tr,i]
    
    y_ALL = np.reshape(y_ALL, (ALL_EEG.size*num_trials, 1)) 


    # # # # # Data augmentation (Down-sampling by 2) # # # # # 
    X_ALL_EEG_DS1 = X_ALL_EEG[:,0::2,:]
    X_ALL_EEG_DS2 = X_ALL_EEG[:,1::2,:]
    X_ALL_EEG_DS  = np.concatenate((X_ALL_EEG_DS1,X_ALL_EEG_DS2))
    # labels
    y_ALL_DS = np.concatenate((y_ALL,y_ALL))

    # # # # # no Shuffle the data # # # # # 
    N = X_ALL_EEG_DS.shape[0]
    indices = [i for i in range(N)]
    #shuffle(indices)
    X_ALL_EEG_DS = X_ALL_EEG_DS[indices, :,:]
    y_ALL_DS     = y_ALL_DS[indices,]

    # Sanity Check: Exclude labels with NaN values
    # Indices of nan and inf values
    idx = np.where((np.isnan(y_ALL_DS)==False) & (np.isinf(y_ALL_DS)==False))
    filtered_X_ALL_EEG_DS = X_ALL_EEG_DS[idx[0],:,:]
    filtered_y_ALL_DS = y_ALL_DS[idx[0]]
    
    
    X = filtered_X_ALL_EEG_DS
    y = filtered_y_ALL_DS
    
    # Make X and y as float 32
    X = X.astype('float32')
    y = y.astype('float32')
    
    # normalize the dataset
    scaler = MinMaxScaler(feature_range=(0, 1))
    
    # Standarizing X
    m, n, d = X.shape
    X = X.reshape(m, n*d)
    X = scaler.fit_transform(X)
    X = X.reshape(m, n, d)
    # Standarizing y
    y = scaler.fit_transform(y)
    

    # make predictions
    y_Predicted = best_model.predict(X)

    
    MSE_Test = mean_squared_error(y, y_Predicted)
    R2_score_Test = r2_score(y, y_Predicted)
    
#    srcfile = openpyxl.load_workbook('Results.xlsm',read_only=False, keep_vba= True)
#
#    sheetname = srcfile.get_sheet_by_name('NewTest_MSE')
#    sheetname.cell(row=PatientID+2,column=ExcCol).value = str(round(MSE_Test,2)) 
#    sheetname = srcfile.get_sheet_by_name('NewTest_R2')
#    sheetname.cell(row=PatientID+2,column=ExcCol).value = str(round(R2_score_Test,2)) 
#
#    srcfile.save('Results.xlsm')
    return (y-y_Predicted)








def headplot(effects):
    pos = [[173, 52],[211, 48],[250 ,50],[86 ,95],[149, 120],[211, 124],[272, 121],[335, 94],[99 ,153],[322, 152],[58 ,203],[133, 203],[210 ,200],[288 ,202],[100 ,246],[320, 250],[406, 221],[86 ,307],[150 ,280],[210, 280],[270 ,280],[337, 306],[142 ,328],[281, 326],[173, 354],[212, 355],[247, 352]]
    Names = ['FP1', 'FPZ', 'FP2','F7', 'F3', 'Fz', 'F4', 'F8', 'Fc5', 'Fc6', 'T7', 'C3', 'Cz', 'C4', 'T8', 'CP6', 'A2', 'P7', 'P3', 'Pz', 'P4', 'P8', 'Po5', 'Po6', 'O1', 'Oz', 'O2']
    mypos = np.array(pos, ndmin=2)
    mv.plot_topomap(effects,mypos, cmap ='RdBu_r', names = Names, show_names = True)
    return (mv)


def Test_Temporal3Parts_6bars_ShortLong (Model, subband, PatientID, part,SL):
    
    best_model = load_model(Model)
    

    if subband == 'full':
       data  = sio.loadmat('shamdata_tlgo')
       HC    = data['shamhceeg'][:,PatientID] 
    else:
       data  = sio.loadmat('DatahcAlireza')
       HC     = data[subband][:,PatientID] # Healthy Control

    # Load the EEG data labels
    labels  = sio.loadmat('task_gvssham')
    HC_labels     = labels['hcoffmed'][:,PatientID]
    
    
    ALL_EEG = HC
    ALL_Beh = HC_labels


    num_chans,time_steps,num_trials = ALL_EEG[0].shape
    X_ALL_EEG = np.zeros((ALL_EEG.size,num_chans,time_steps,num_trials))
    
    
    x = ALL_EEG[0]
    X_ALL_EEG[:,:,:] = x
    # Data transpose     
    X_ALL_EEG = np.transpose(X_ALL_EEG, (0, 3, 2, 1))        
    # Data dimensions          
    p,m,n,d = X_ALL_EEG.shape
    # Data reshape     
    X_ALL_EEG =  X_ALL_EEG.reshape(p* m,n,d)  
    X2_ALL_EEG =  np.zeros((p* m,n,d)) 

    # # # # # Load the EEG data labels # # # # # 
    num_trials, num_behvars = ALL_Beh[0].shape
    movement_time_ALL = np.zeros((ALL_EEG.size, num_trials,1)) 
    reaction_time_ALL = np.zeros((ALL_EEG.size, num_trials,1))                

    y_ALL = np.zeros((ALL_EEG.size, num_trials,1)) 
    ALLBeh = ALL_Beh[0]
    for tr in range (num_trials):
        i = 12 # index of Reaction Time (Change the index to look other Behavior measures)
        y_ALL[0,tr] = ALLBeh[tr,i]
        
        
    for tr in range (num_trials):
        i = 6 #movement time
        movement_time_ALL[0,tr] = ALLBeh[tr,i]
        
        i = 12
        reaction_time_ALL[0,tr] = ALLBeh[tr,i]

    rt_df = pd.DataFrame(reaction_time_ALL[0])
    rt_df.fillna((rt_df.mean()), inplace=True) 
    rt_df.to_numpy()
    mt_df = pd.DataFrame(movement_time_ALL[0])
    mt_df.fillna((mt_df.mean()), inplace=True)
    mt_df.to_numpy()
    
    
    max_RT_ind = rt_df.idxmax()
    max_RT     = rt_df.max()
    min_RT_ind = rt_df.idxmin()
    min_RT     = rt_df.min()

    if SL == 'S':
        myRT = min_RT_ind
    else:
        myRT = max_RT_ind
        
    tr = myRT
    if part == 1:
            t1 = 0
            t2 = 1000
            temp_con = np.concatenate((np.zeros(t1), np.ones(t2-t1), np.zeros(2000-t2)), axis=0)
    elif part == 2:
            t1 = 1000
            t2 = int(1000 + rt_df[0][tr])
            temp_con = np.concatenate((np.zeros(t1), np.ones(t2-t1), np.zeros(2000-t2)), axis=0)
    else:
            t1 = int(1000 + rt_df[0][tr])
            t2 = 1999
            temp_con = np.concatenate((np.zeros(t1), np.ones(t2-t1), np.zeros(2000-t2)), axis=0)
                  
        
            
    repetitions = 27
    repeats_array = np.transpose(np.tile(temp_con, (repetitions,1)))
    temp = X_ALL_EEG[tr,:,:]
    X2_ALL_EEG[tr,:,:] = temp * repeats_array

    
    y_ALL = np.reshape(y_ALL, (ALL_EEG.size*num_trials, 1)) 


    # # # # # Data augmentation (Down-sampling by 2) # # # # # 
    X_ALL_EEG_DS1 = X2_ALL_EEG[:,0::2,:]
    X_ALL_EEG_DS2 = X2_ALL_EEG[:,1::2,:]
    X_ALL_EEG_DS  = np.concatenate((X_ALL_EEG_DS1,X_ALL_EEG_DS2))
    # labels
    y_ALL_DS = np.concatenate((y_ALL,y_ALL))

    # # # # # no Shuffle the data # # # # # 
    N = X_ALL_EEG_DS.shape[0]
    indices = [i for i in range(N)]
    #shuffle(indices)
    X_ALL_EEG_DS = X_ALL_EEG_DS[indices, :,:]
    y_ALL_DS     = y_ALL_DS[indices,]

    # Sanity Check: Exclude labels with NaN values
    # Indices of nan and inf values
    idx = np.where((np.isnan(y_ALL_DS)==False) & (np.isinf(y_ALL_DS)==False))
    filtered_X_ALL_EEG_DS = X_ALL_EEG_DS[idx[0],:,:]
    filtered_y_ALL_DS = y_ALL_DS[idx[0]]
    
    
    X = filtered_X_ALL_EEG_DS
    y = filtered_y_ALL_DS
    
    # Make X and y as float 32
    X = X.astype('float32')
    y = y.astype('float32')
    
    # normalize the dataset
    scaler = MinMaxScaler(feature_range=(0, 1))
    
    # Standarizing X
    m, n, d = X.shape
    X = X.reshape(m, n*d)
    X = scaler.fit_transform(X)
    X = X.reshape(m, n, d)
    # Standarizing y
    y = scaler.fit_transform(y)
    

    # make predictions
    y_Predicted = best_model.predict(X)

    MSE_Test = mean_squared_error(y, y_Predicted)
    R2_score_Test = r2_score(y, y_Predicted)
    #return (y-y_Predicted)    
    return(MSE_Test)